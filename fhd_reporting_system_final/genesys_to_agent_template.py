import argparse
import datetime as dt
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description='Map Genesys Agent Performance + Status (+ Booking) into Agent Report Template')
    p.add_argument('--perf', required=True, help='Genesys Agent Performance Summary CSV')
    p.add_argument('--status', required=True, help='Genesys Agent Status Summary CSV')
    p.add_argument('--booking', required=False, help='Booking CSV file (Salamtak)')
    p.add_argument('--template', required=True, help='Excel template: Agent Report Template.xlsx')
    p.add_argument('--out-xlsx', default='Agent_Productivity_Filled.xlsx', help='Output filled Excel file')
    p.add_argument('--out-csv', default='Agent_Productivity_Filled.csv', help='Output aggregated CSV file')
    p.add_argument('--encoding', default='utf-8-sig', help='CSV encoding for Genesys files')
    return p.parse_args()


def parse_interval_date(val: str) -> dt.date:
    if not isinstance(val, str):
        return dt.date.today()
    for fmt in ('%d/%m/%y %H:%M', '%m/%d/%y %H:%M'):
        try:
            return dt.datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return dt.date.today()


def seconds_to_excel_time(seconds: Optional[float]) -> Optional[float]:
    if seconds is None:
        return None
    try:
        s = float(seconds)
    except (TypeError, ValueError):
        return None
    if pd.isna(s):
        return None
    return s / 86400.0


def seconds_to_hhmmss_time(seconds: Optional[float]) -> Optional[dt.time]:
    if seconds is None:
        return None
    try:
        s = int(round(float(seconds)))
    except (TypeError, ValueError):
        return None
    if s < 0:
        s = 0
    h = s // 3600
    m = (s % 3600) // 60
    sec = s % 60
    if h > 23:
        h, m, sec = 23, 59, 59
    return dt.time(hour=h, minute=m, second=sec)


def load_perf(perf_csv: Path, encoding: str) -> pd.DataFrame:
    df = pd.read_csv(perf_csv, encoding=encoding)
    if 'Media Type' in df.columns:
        df = df[df['Media Type'].fillna('') == 'voice']
    return df


def load_status(status_csv: Path, encoding: str) -> pd.DataFrame:
    df = pd.read_csv(status_csv, encoding=encoding)
    return df


def load_booking(booking_csv: Path) -> pd.DataFrame:
    df = pd.read_csv(booking_csv, encoding='utf-8', encoding_errors='replace')
    return df


def aggregate_perf(df: pd.DataFrame) -> pd.DataFrame:
    group_cols = []
    for col in ['Interval Start', 'Agent Id', 'Agent Name', 'Division Name']:
        if col in df.columns:
            group_cols.append(col)

    agg = {}
    if 'Answered' in df.columns:
        agg['Answered'] = 'sum'
    if 'Outbound' in df.columns:
        agg['Outbound'] = 'sum'
    if 'Total ACW' in df.columns:
        agg['Total ACW'] = 'sum'
    if 'Avg Handle' in df.columns:
        agg['Avg Handle'] = 'mean'
    if 'Total Handle' in df.columns:
        agg['Total Handle'] = 'sum'

    grouped = df.groupby(group_cols, dropna=False).agg(agg).reset_index()
    return grouped


def aggregate_status(df: pd.DataFrame) -> pd.DataFrame:
    group_cols = []
    for col in ['Interval Start', 'Agent Id', 'Agent Name', 'Division Name']:
        if col in df.columns:
            group_cols.append(col)

    agg = {}
    for col in ['Logged In']:
        if col in df.columns:
            agg[col] = 'sum'
    for col in ['Log In', 'Log Out']:
        if col in df.columns:
            agg[col] = 'first'

    grouped = df.groupby(group_cols, dropna=False).agg(agg).reset_index()
    return grouped


def aggregate_booking(df: pd.DataFrame) -> pd.DataFrame:
    name_col = None
    for c in df.columns:
        key = str(c).strip().upper()
        if key in ('CC_CLERK_NAME', 'AGENT_NAME', 'AGENT NAME'):
            name_col = c
            break
    if name_col is None:
        raise ValueError('Booking file does not contain CC_CLERK_NAME / AGENT_NAME column')

    count_col = None
    for c in df.columns:
        key = str(c).upper()
        if 'NO_OF_BOOKED_APPT' in key or 'BOOKED' in key:
            count_col = c
            break
    if count_col is None:
        raise ValueError('Booking file does not contain NO_OF_BOOKED_APPT-like column')

    df2 = df.copy()
    df2['Agent Name'] = df2[name_col].astype(str)
    df2['Total Inbound Booking'] = pd.to_numeric(df2[count_col], errors='coerce').fillna(0.0)

    grouped = df2.groupby('Agent Name', dropna=False)['Total Inbound Booking'].sum().reset_index()
    return grouped


def merge_perf_status(perf_agg: pd.DataFrame, status_agg: pd.DataFrame) -> pd.DataFrame:
    join_cols = [c for c in ['Interval Start', 'Agent Id', 'Agent Name', 'Division Name']
                 if c in perf_agg.columns and c in status_agg.columns]
    if not join_cols:
        join_cols = [c for c in ['Agent Id', 'Agent Name']
                     if c in perf_agg.columns and c in status_agg.columns]

    merged = pd.merge(perf_agg, status_agg, on=join_cols, how='outer', suffixes=('_perf', '_status'))
    return merged


def fill_template(template_path: Path, df: pd.DataFrame, out_xlsx: Path, out_csv: Path) -> None:
    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb['Sheet1']

    max_row = ws.max_row
    for r in range(3, max_row + 1):
        for c in range(1, 15):
            ws.cell(row=r, column=c).value = None

    records = []
    row_idx = 3

    for _, row in df.iterrows():
        agent_name = row.get('Agent Name', '')
        division = row.get('Division Name', '')

        interval_start = row.get('Interval Start')
        if isinstance(interval_start, str):
            date_val = parse_interval_date(interval_start)
        elif isinstance(interval_start, dt.datetime):
            date_val = interval_start.date()
        else:
            date_val = dt.date.today()

        log_in_raw = row.get('Log In')
        log_out_raw = row.get('Log Out')

        logged_in_seconds = row.get('Logged In')
        logged_in_duration = seconds_to_excel_time(logged_in_seconds)

        raw_answered = row.get('Answered')
        answered = 0 if pd.isna(raw_answered) else float(raw_answered)

        raw_outbound = row.get('Outbound')
        outbound = 0 if pd.isna(raw_outbound) else float(raw_outbound)

        inq_amb = answered

        inbound_booking_val = row.get('Total Inbound Booking', 0)
        inbound_booking = 0 if pd.isna(inbound_booking_val) else float(inbound_booking_val)

        raw_acw = row.get('Total ACW')
        total_acw_sec = 0 if pd.isna(raw_acw) else float(raw_acw)
        total_acw_excel = seconds_to_excel_time(total_acw_sec)

        raw_total_handle = row.get('Total Handle')
        total_handle_sec = None if pd.isna(raw_total_handle) else float(raw_total_handle)

        if total_handle_sec and answered:
            avg_handle_sec = float(total_handle_sec) / float(answered)
        else:
            raw_avg_handle = row.get('Avg Handle')
            avg_handle_sec = None if pd.isna(raw_avg_handle) else float(raw_avg_handle)

        avg_handle_time = seconds_to_hhmmss_time(avg_handle_sec)

        ws[f'A{row_idx}'] = dt.datetime.combine(date_val, dt.time(6, 30))
        ws[f'B{row_idx}'] = agent_name
        ws[f'C{row_idx}'] = division
        ws[f'D{row_idx}'] = log_in_raw
        ws[f'E{row_idx}'] = log_out_raw

        if logged_in_duration is not None:
            ws[f'F{row_idx}'] = logged_in_duration

        ws[f'G{row_idx}'] = int(answered)
        ws[f'H{row_idx}'] = int(inq_amb)
        ws[f'I{row_idx}'] = int(outbound)
        ws[f'J{row_idx}'] = int(inbound_booking)
        ws[f'K{row_idx}'] = f"=MAX(0,IFERROR(J{row_idx}/(G{row_idx}-H{row_idx}),0))"
        ws[f'L{row_idx}'] = f"=G{row_idx}+I{row_idx}"

        if total_acw_excel is not None:
            ws[f'M{row_idx}'] = total_acw_excel
        if avg_handle_time is not None:
            ws[f'N{row_idx}'] = avg_handle_time

        records.append({
            'Date': date_val.isoformat(),
            'Agent Name': agent_name,
            'Division Name': division,
            'Log in Time': log_in_raw,
            'Log Out Time': log_out_raw,
            'Total Logged In Duration (sec)': logged_in_seconds,
            'Total Answered Calls': answered,
            'Answered INQ&AMB Queue Calls': inq_amb,
            'Outbound Calls': outbound,
            'Total Inbound Booking': inbound_booking,
            'Total Wrap-Up (ACW) sec': total_acw_sec,
            'Avg Handle Time sec': avg_handle_sec,
        })

        row_idx += 1

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame.from_records(records).to_csv(out_csv, index=False, encoding='utf-8-sig')


def run_from_paths(
    perf: Path,
    status: Path,
    template: Path,
    out_xlsx: Path,
    out_csv: Path,
    encoding: str = 'utf-8-sig',
    booking: Optional[Path] = None,
) -> None:
    perf_df = load_perf(perf, encoding)
    status_df = load_status(status, encoding)
    perf_agg = aggregate_perf(perf_df)
    status_agg = aggregate_status(status_df)
    merged = merge_perf_status(perf_agg, status_agg)

    if booking is not None and booking.exists():
        booking_df = load_booking(booking)
        booking_agg = aggregate_booking(booking_df)
        if 'Agent Name' in merged.columns and 'Agent Name' in booking_agg.columns:
            merged = pd.merge(merged, booking_agg, on='Agent Name', how='left')

    if 'Agent Name' in merged.columns:
        merged = merged[merged['Agent Name'].notna()]

    fill_template(template, merged, out_xlsx, out_csv)


def main():
    args = parse_args()
    perf_path = Path(args.perf)
    status_path = Path(args.status)
    template_path = Path(args.template)
    out_xlsx = Path(args.out_xlsx)
    out_csv = Path(args.out_csv)
    booking_path = Path(args.booking) if args.booking else None

    run_from_paths(perf_path, status_path, template_path, out_xlsx, out_csv, args.encoding, booking_path)


if __name__ == '__main__':
    main()